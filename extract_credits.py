import openpyxl
import json

wb = openpyxl.load_workbook('ref/pbrs credits.xlsx')

# Extract data from each sheet
checkbox_credits = {}
dropdown_credits = {}
straight_credits = {}

# CHECKBOX SHEET
ws = wb['Checkbox']
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] is None or row[0] == 'Credit':
        continue
    credit_code = str(row[0]).strip()
    pts = row[1]
    desc = str(row[2]).strip() if row[2] else ""
    
    if credit_code not in checkbox_credits:
        checkbox_credits[credit_code] = {'pts': pts, 'options': []}
    checkbox_credits[credit_code]['options'].append({'pts': pts, 'desc': desc})

# DROPDOWN SHEET
ws = wb['Dropdown']
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] is None or row[0] == 'Credit':
        continue
    credit_code = str(row[0]).strip()
    pts = row[1]
    desc = str(row[2]).strip() if row[2] else ""
    
    if credit_code not in dropdown_credits:
        dropdown_credits[credit_code] = {'pts': pts, 'options': []}
    dropdown_credits[credit_code]['options'].append({'pts': pts, 'desc': desc})

# STRAIGHT SHEET
ws = wb['Straight']
for row in ws.iter_rows(min_row=1, values_only=True):
    if row[0] is None or row[0] == 'Credit':
        continue
    credit_code = str(row[0]).strip()
    pts = row[1]
    desc = str(row[2]).strip() if row[2] else ""
    
    straight_credits[credit_code] = {'pts': pts, 'desc': desc}

print("=== CHECKBOX SHEET ===")
print(json.dumps(checkbox_credits, indent=2, default=str)[:2000])

print("\n=== DROPDOWN SHEET (first 5) ===")
dd_list = list(dropdown_credits.items())[:5]
for code, data in dd_list:
    print(f"{code}: {data}")

print("\n=== STRAIGHT SHEET (first 5) ===")
st_list = list(straight_credits.items())[:5]
for code, data in st_list:
    print(f"{code}: {data}")

print(f"\n\nSummary:")
print(f"Checkbox credits: {len(checkbox_credits)}")
print(f"Dropdown credits: {len(dropdown_credits)}")
print(f"Straight credits: {len(straight_credits)}")

# Output for JS file
output = {
    'checkbox': checkbox_credits,
    'dropdown': dropdown_credits,
    'straight': straight_credits
}

with open('credit_categories.json', 'w') as f:
    json.dump(output, f, indent=2, default=str)

print("\nGenerated credit_categories.json")
