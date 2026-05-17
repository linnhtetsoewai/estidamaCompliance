import openpyxl

wb = openpyxl.load_workbook('ref/pbrs credits.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"="*100}')
    print(f'SHEET: {sheet_name.upper()}')
    print(f'{"="*100}')
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        if row[0] is not None:
            col1 = str(row[0])[:50] if row[0] else ""
            col2 = str(row[1]) if row[1] else ""
            col3 = str(row[2])[:50] if len(row) > 2 and row[2] else ""
            print(f"Row {idx}: {col1:<50} | {col2:<10} | {col3}")
